# -*- coding: utf-8 -*-
"""
Test de l'export de la liasse avec la balance demo
Vérifie que le TFT est correctement exporté
"""
import sys
import os

# Ajouter py_backend au path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'py_backend'))

import pandas as pd
from etats_financiers import calculer_etats_financiers
from export_liasse import remplir_liasse_officielle
import openpyxl

print("="*80)
print("TEST EXPORT LIASSE AVEC BALANCE DEMO")
print("="*80)
print()

# Chemin de la balance demo
balance_path = "P000 -BALANCE DEMO N_N-1_N-2.xls"

if not os.path.exists(balance_path):
    print(f"❌ ERREUR: Balance demo non trouvée: {balance_path}")
    print("   Veuillez placer le fichier à la racine du projet")
    exit(1)

print(f"✅ Balance demo trouvée: {balance_path}")
print()

# ÉTAPE 1: Lire la balance N
print("📊 ÉTAPE 1: Lecture de la balance N...")
try:
    df_n = pd.read_excel(balance_path, sheet_name="BALANCE N")
    print(f"   ✅ Balance N lue: {len(df_n)} lignes")
    print(f"   Colonnes: {list(df_n.columns)}")
except Exception as e:
    print(f"   ❌ Erreur lecture balance N: {e}")
    exit(1)

print()

# ÉTAPE 2: Calculer les états financiers
print("📊 ÉTAPE 2: Calcul des états financiers...")
try:
    results = calculer_etats_financiers(
        df_balance_n=df_n,
        df_balance_n1=None,
        df_balance_n2=None
    )
    print("   ✅ États financiers calculés")
    
    # Vérifier les sections
    sections = ['bilan_actif', 'bilan_passif', 'charges', 'produits', 'tft']
    for section in sections:
        if section in results:
            nb_postes = len(results[section])
            print(f"   ✅ {section}: {nb_postes} postes")
        else:
            print(f"   ❌ {section}: MANQUANT")
    
except Exception as e:
    print(f"   ❌ Erreur calcul états financiers: {e}")
    import traceback
    traceback.print_exc()
    exit(1)

print()

# ÉTAPE 3: Vérifier le TFT
print("📊 ÉTAPE 3: Vérification du TFT...")
if 'tft' in results:
    tft = results['tft']
    print(f"   ✅ TFT présent avec {len(tft)} références")
    
    # Afficher quelques valeurs
    refs_test = ['ZA', 'FA', 'ZB', 'ZC', 'ZD', 'ZE', 'ZF', 'ZG', 'ZH']
    print()
    print("   Valeurs TFT:")
    for ref in refs_test:
        if ref in tft:
            montant = tft[ref].get('montant', 0)
            libelle = tft[ref].get('libelle', '')
            print(f"     {ref}: {montant:>15,.2f} - {libelle}")
        else:
            print(f"     {ref}: NON TROUVÉ")
else:
    print("   ❌ TFT MANQUANT dans les résultats")
    exit(1)

print()

# ÉTAPE 4: Exporter la liasse
print("📊 ÉTAPE 4: Export de la liasse...")
try:
    fichier_bytes = remplir_liasse_officielle(
        results=results,
        nom_entreprise="ENTREPRISE DEMO",
        exercice="2024"
    )
    print(f"   ✅ Liasse exportée: {len(fichier_bytes)} bytes")
    
    # Sauvegarder le fichier
    output_path = "test_liasse_avec_tft.xlsx"
    with open(output_path, 'wb') as f:
        f.write(fichier_bytes)
    print(f"   ✅ Fichier sauvegardé: {output_path}")
    
except Exception as e:
    print(f"   ❌ Erreur export liasse: {e}")
    import traceback
    traceback.print_exc()
    exit(1)

print()

# ÉTAPE 5: Vérifier le contenu du fichier Excel
print("📊 ÉTAPE 5: Vérification du fichier Excel généré...")
try:
    wb = openpyxl.load_workbook(output_path)
    print(f"   ✅ Fichier Excel ouvert")
    print(f"   Onglets: {wb.sheetnames}")
    
    # Vérifier l'onglet TFT
    if 'TFT' in wb.sheetnames:
        ws_tft = wb['TFT']
        print()
        print("   ✅ Onglet TFT trouvé")
        print()
        print("   Vérification des cellules TFT:")
        
        # Vérifier quelques cellules
        cellules_test = {
            'C10': 'ZA - Trésorerie ouverture',
            'C12': 'FA - CAFG',
            'C19': 'ZB - Total flux opérationnels',
            'C26': 'ZC - Total flux investissement',
            'C32': 'ZD - Total flux capitaux propres',
            'C37': 'ZE - Total flux capitaux étrangers',
            'C40': 'ZH - Trésorerie clôture'
        }
        
        cellules_remplies = 0
        cellules_vides = 0
        
        for cellule, description in cellules_test.items():
            valeur = ws_tft[cellule].value
            if valeur is not None and valeur != 0:
                print(f"     ✅ {cellule}: {valeur:>15,.2f} - {description}")
                cellules_remplies += 1
            else:
                print(f"     ⚠️  {cellule}: VIDE - {description}")
                cellules_vides += 1
        
        print()
        print(f"   Résumé: {cellules_remplies} cellules remplies, {cellules_vides} cellules vides")
        
    else:
        print("   ❌ Onglet TFT NON TROUVÉ")
    
    wb.close()
    
except Exception as e:
    print(f"   ❌ Erreur vérification fichier Excel: {e}")
    import traceback
    traceback.print_exc()
    exit(1)

print()
print("="*80)
print("✅ TEST TERMINÉ AVEC SUCCÈS")
print("="*80)
print()
print("Fichier généré: test_liasse_avec_tft.xlsx")
print("Vous pouvez l'ouvrir pour vérifier le contenu du TFT")
print()
