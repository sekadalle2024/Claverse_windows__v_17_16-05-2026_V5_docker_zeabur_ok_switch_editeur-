# -*- coding: utf-8 -*-
"""
Script pour vérifier les cellules du TFT dans le template
"""
import openpyxl

wb = openpyxl.load_workbook('py_backend/Liasse_officielle_revise.xlsx')
ws = wb['TFT']

print("="*80)
print("VÉRIFICATION DES CELLULES TFT DANS LE TEMPLATE")
print("="*80)
print()

# Vérifier quelques cellules clés
cellules_test = {
    'C10': 'ZA - Trésorerie ouverture',
    'C12': 'FA - CAFG',
    'C17': 'ZB - Total flux opérationnels',
    'C24': 'ZC - Total flux investissement',
    'C30': 'ZD - Total flux capitaux propres',
    'C35': 'ZE - Total flux capitaux étrangers',
    'C38': 'ZH - Trésorerie clôture'
}

print("Cellules mappées:")
for cellule, description in cellules_test.items():
    valeur = ws[cellule].value
    print(f"  {cellule}: {description}")
    print(f"    Valeur actuelle: {valeur}")
    print()

# Vérifier la colonne A pour les références
print("Références dans la colonne A:")
for row in range(10, 40):
    ref = ws[f'A{row}'].value
    if ref:
        print(f"  A{row}: {ref}")

wb.close()
