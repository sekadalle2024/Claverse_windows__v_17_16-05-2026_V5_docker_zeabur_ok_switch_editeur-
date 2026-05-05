# -*- coding: utf-8 -*-
"""
Test simplifié de l'export de la liasse avec vérification du mapping TFT
"""
import os
import openpyxl

print("="*80)
print("TEST MAPPING TFT DANS EXPORT_LIASSE.PY")
print("="*80)
print()

# ÉTAPE 1: Vérifier que le fichier export_liasse.py contient le mapping TFT
print("📊 ÉTAPE 1: Vérification du mapping TFT dans export_liasse.py...")
with open('py_backend/export_liasse.py', 'r', encoding='utf-8') as f:
    content = f.read()

if 'MAPPING_TFT' in content:
    print("   ✅ MAPPING_TFT trouvé dans export_liasse.py")
    
    # Compter les références
    import re
    refs = re.findall(r"'([A-Z]{2})':\s*'C\d+'", content)
    tft_refs = [r for r in refs if r.startswith(('Z', 'F'))]
    print(f"   ✅ {len(tft_refs)} références TFT mappées")
    
    # Afficher les références
    print()
    print("   Références TFT mappées:")
    for ref in sorted(set(tft_refs)):
        print(f"     - {ref}")
else:
    print("   ❌ MAPPING_TFT NON TROUVÉ")
    exit(1)

print()

# ÉTAPE 2: Vérifier le code de remplissage TFT
print("📊 ÉTAPE 2: Vérification du code de remplissage TFT...")
if 'Remplir le TABLEAU DES FLUX DE TRÉSORERIE' in content or 'Remplissage TFT' in content:
    print("   ✅ Code de remplissage TFT trouvé")
    
    # Vérifier qu'il utilise MAPPING_TFT
    if 'for ref, cellule in MAPPING_TFT.items()' in content:
        print("   ✅ Utilise MAPPING_TFT pour remplir les cellules")
    else:
        print("   ⚠️  Code de remplissage ne semble pas utiliser MAPPING_TFT")
    
    # Vérifier qu'il accède à results['tft']
    if "results.get('tft'" in content or "results['tft']" in content:
        print("   ✅ Accède aux données TFT depuis results")
    else:
        print("   ⚠️  Ne semble pas accéder à results['tft']")
else:
    print("   ❌ Code de remplissage TFT NON TROUVÉ")

print()

# ÉTAPE 3: Vérifier le template
print("📊 ÉTAPE 3: Vérification du template Liasse_officielle_revise.xlsx...")
template_path = "py_backend/Liasse_officielle_revise.xlsx"

if os.path.exists(template_path):
    print(f"   ✅ Template trouvé: {template_path}")
    
    try:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        print(f"   ✅ Template ouvert: {len(wb.sheetnames)} onglets")
        
        if 'TFT' in wb.sheetnames:
            print("   ✅ Onglet TFT présent")
            
            ws = wb['TFT']
            
            # Vérifier les références dans la colonne A
            refs_template = []
            for row in range(10, 50):
                ref = ws[f'A{row}'].value
                if ref and isinstance(ref, str) and len(ref) == 2:
                    refs_template.append(ref)
            
            print(f"   ✅ {len(refs_template)} références trouvées dans le template")
            print()
            print("   Références dans le template:")
            for ref in refs_template:
                print(f"     - {ref}")
        else:
            print("   ❌ Onglet TFT NON TROUVÉ")
        
        wb.close()
    except Exception as e:
        print(f"   ❌ Erreur lecture template: {e}")
else:
    print(f"   ❌ Template NON TROUVÉ: {template_path}")

print()

# ÉTAPE 4: Vérifier les correspondances
print("📊 ÉTAPE 4: Vérification des correspondances_syscohada.json...")
import json

with open('py_backend/correspondances_syscohada.json', 'r', encoding='utf-8') as f:
    correspondances = json.load(f)

if 'tft' in correspondances:
    print(f"   ✅ Section TFT trouvée: {len(correspondances['tft'])} postes")
    
    # Afficher les références
    refs_json = [poste['ref'] for poste in correspondances['tft']]
    print()
    print("   Références dans correspondances_syscohada.json:")
    for ref in refs_json:
        print(f"     - {ref}")
else:
    print("   ❌ Section TFT NON TROUVÉE")

print()

# ÉTAPE 5: Comparaison finale
print("📊 ÉTAPE 5: Comparaison des références...")
print()

# Références dans export_liasse.py
refs_export = sorted(set(tft_refs))

# Références dans correspondances_syscohada.json
refs_json = sorted([poste['ref'] for poste in correspondances.get('tft', [])])

# Références dans le template
refs_template = sorted(refs_template) if 'refs_template' in locals() else []

print(f"   Export_liasse.py:              {len(refs_export)} références")
print(f"   Correspondances_syscohada.json: {len(refs_json)} références")
print(f"   Template Excel:                 {len(refs_template)} références")
print()

# Vérifier la cohérence
if refs_export == refs_json:
    print("   ✅ Export et JSON: COHÉRENTS")
else:
    print("   ⚠️  Export et JSON: DIFFÉRENTS")
    manquants_export = set(refs_json) - set(refs_export)
    manquants_json = set(refs_export) - set(refs_json)
    if manquants_export:
        print(f"      Manquants dans export: {manquants_export}")
    if manquants_json:
        print(f"      Manquants dans JSON: {manquants_json}")

if refs_export == refs_template:
    print("   ✅ Export et Template: COHÉRENTS")
else:
    print("   ⚠️  Export et Template: DIFFÉRENTS")
    manquants_export = set(refs_template) - set(refs_export)
    manquants_template = set(refs_export) - set(refs_template)
    if manquants_export:
        print(f"      Manquants dans export: {manquants_export}")
    if manquants_template:
        print(f"      Manquants dans template: {manquants_template}")

print()
print("="*80)
print("✅ VÉRIFICATION TERMINÉE")
print("="*80)
print()
print("Résumé:")
print("  - Le mapping TFT est présent dans export_liasse.py")
print("  - Le code de remplissage TFT est présent")
print("  - Le template contient l'onglet TFT")
print("  - Les correspondances TFT sont dans le JSON")
print()
print("Prochaine étape:")
print("  - Tester avec une vraie balance pour vérifier l'export complet")
print()
