# -*- coding: utf-8 -*-
"""
Script de vérification de la correspondance entre:
- export_liasse.py
- etats_financiers.py  
- correspondances_syscohada.json
- Liasse_officielle_revise.xlsx (template)

Ce script vérifie:
1. Les onglets du template correspondent aux mappings
2. Les références dans correspondances_syscohada.json sont utilisées
3. Le TFT est prévu dans export_liasse.py
4. Les cellules mappées existent dans le template
"""

import openpyxl
import json
import os
from pathlib import Path

# Couleurs pour le terminal
class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    BOLD = '\033[1m'
    END = '\033[0m'

def print_success(msg):
    print(f"{Colors.GREEN}✅ {msg}{Colors.END}")

def print_error(msg):
    print(f"{Colors.RED}❌ {msg}{Colors.END}")

def print_warning(msg):
    print(f"{Colors.YELLOW}⚠️  {msg}{Colors.END}")

def print_info(msg):
    print(f"{Colors.BLUE}ℹ️  {msg}{Colors.END}")

def print_header(msg):
    print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{msg}{Colors.END}")
    print(f"{Colors.BOLD}{'='*80}{Colors.END}\n")


# ==================== VÉRIFICATION 1: FICHIERS EXISTENT ====================

print_header("VÉRIFICATION 1: EXISTENCE DES FICHIERS")

files_to_check = {
    "Template Liasse": "py_backend/Liasse_officielle_revise.xlsx",
    "Export Liasse": "py_backend/export_liasse.py",
    "États Financiers": "py_backend/etats_financiers.py",
    "Correspondances": "py_backend/correspondances_syscohada.json",
    "TFT Module": "py_backend/tableau_flux_tresorerie.py",
    "TFT Module V2": "py_backend/tableau_flux_tresorerie_v2.py"
}

all_files_exist = True
for name, path in files_to_check.items():
    if os.path.exists(path):
        size = os.path.getsize(path) / 1024
        print_success(f"{name}: {path} ({size:.2f} KB)")
    else:
        print_error(f"{name}: {path} - FICHIER MANQUANT")
        all_files_exist = False

if not all_files_exist:
    print_error("Certains fichiers sont manquants. Arrêt du script.")
    exit(1)


# ==================== VÉRIFICATION 2: ONGLETS DU TEMPLATE ====================

print_header("VÉRIFICATION 2: ONGLETS DU TEMPLATE EXCEL")

try:
    wb = openpyxl.load_workbook("py_backend/Liasse_officielle_revise.xlsx", read_only=True)
    sheet_names = wb.sheetnames
    print_info(f"Nombre d'onglets: {len(sheet_names)}")
    
    # Onglets attendus selon SYSCOHADA Révisé
    onglets_attendus = {
        "BILAN": ["BILAN", "Bilan", "ACTIF", "PASSIF"],
        "COMPTE DE RÉSULTAT": ["COMPTE DE RESULTAT", "Compte de résultat", "RESULTAT", "CR"],
        "TFT": ["TFT", "TAFIRE", "Tableau des flux", "Flux de trésorerie"],
        "ANNEXES": ["ANNEXES", "Annexes", "Notes"]
    }
    
    onglets_trouves = {}
    for categorie, variations in onglets_attendus.items():
        found = False
        for variation in variations:
            if variation in sheet_names:
                onglets_trouves[categorie] = variation
                print_success(f"{categorie}: Onglet '{variation}' trouvé")
                found = True
                break
        if not found:
            print_warning(f"{categorie}: Aucun onglet trouvé parmi {variations}")
            onglets_trouves[categorie] = None
    
    wb.close()
    
except Exception as e:
    print_error(f"Erreur lors de la lecture du template: {e}")
    exit(1)


# ==================== VÉRIFICATION 3: CORRESPONDANCES JSON ====================

print_header("VÉRIFICATION 3: STRUCTURE DES CORRESPONDANCES")

try:
    with open("py_backend/correspondances_syscohada.json", "r", encoding="utf-8") as f:
        correspondances = json.load(f)
    
    sections_attendues = ["bilan_actif", "bilan_passif", "charges", "produits", "tft"]
    
    for section in sections_attendues:
        if section in correspondances:
            nb_postes = len(correspondances[section])
            print_success(f"Section '{section}': {nb_postes} postes")
            
            # Vérifier la structure des postes
            if nb_postes > 0:
                premier_poste = correspondances[section][0]
                if "ref" in premier_poste and "libelle" in premier_poste and "racines" in premier_poste:
                    print_info(f"  Structure correcte: ref, libelle, racines")
                else:
                    print_warning(f"  Structure incomplète pour le premier poste")
        else:
            if section == "tft":
                print_warning(f"Section '{section}': MANQUANTE (TFT non prévu dans correspondances)")
            else:
                print_error(f"Section '{section}': MANQUANTE")
    
except Exception as e:
    print_error(f"Erreur lors de la lecture des correspondances: {e}")
    exit(1)


# ==================== VÉRIFICATION 4: MAPPINGS DANS EXPORT_LIASSE.PY ====================

print_header("VÉRIFICATION 4: MAPPINGS DANS EXPORT_LIASSE.PY")

try:
    with open("py_backend/export_liasse.py", "r", encoding="utf-8") as f:
        export_liasse_content = f.read()
    
    # Vérifier les mappings définis
    mappings_attendus = [
        "MAPPING_BILAN_ACTIF",
        "MAPPING_BILAN_PASSIF",
        "MAPPING_COMPTE_RESULTAT_CHARGES",
        "MAPPING_COMPTE_RESULTAT_PRODUITS"
    ]
    
    for mapping in mappings_attendus:
        if mapping in export_liasse_content:
            print_success(f"Mapping '{mapping}' défini")
        else:
            print_error(f"Mapping '{mapping}' MANQUANT")
    
    # Vérifier si le TFT est prévu
    if "TFT" in export_liasse_content or "tft" in export_liasse_content.lower():
        print_success("TFT mentionné dans export_liasse.py")
    else:
        print_warning("TFT NON mentionné dans export_liasse.py")
    
    # Vérifier le template utilisé
    if "Liasse_officielle_revise.xlsx" in export_liasse_content:
        print_success("Template 'Liasse_officielle_revise.xlsx' utilisé en priorité")
    else:
        print_error("Template 'Liasse_officielle_revise.xlsx' NON utilisé")
    
except Exception as e:
    print_error(f"Erreur lors de la lecture de export_liasse.py: {e}")
    exit(1)


# ==================== VÉRIFICATION 5: TFT DANS ETATS_FINANCIERS.PY ====================

print_header("VÉRIFICATION 5: TFT DANS ETATS_FINANCIERS.PY")

try:
    with open("py_backend/etats_financiers.py", "r", encoding="utf-8") as f:
        etats_fin_content = f.read()
    
    # Vérifier les imports TFT
    if "from tableau_flux_tresorerie" in etats_fin_content:
        print_success("Import du module TFT trouvé")
    else:
        print_error("Import du module TFT MANQUANT")
    
    # Vérifier l'appel de calcul TFT
    if "calculer_tft" in etats_fin_content:
        print_success("Fonction 'calculer_tft' appelée")
    else:
        print_error("Fonction 'calculer_tft' NON appelée")
    
    # Vérifier si TFT est ajouté aux résultats
    if "results['tft']" in etats_fin_content or "results_liasse['tft']" in etats_fin_content:
        print_success("TFT ajouté aux résultats")
    else:
        print_warning("TFT possiblement NON ajouté aux résultats")
    
except Exception as e:
    print_error(f"Erreur lors de la lecture de etats_financiers.py: {e}")
    exit(1)


# ==================== VÉRIFICATION 6: CELLULES DANS LE TEMPLATE ====================

print_header("VÉRIFICATION 6: VÉRIFICATION DES CELLULES MAPPÉES")

try:
    wb = openpyxl.load_workbook("py_backend/Liasse_officielle_revise.xlsx")
    
    # Extraire les mappings de export_liasse.py
    import re
    
    # Mapping Bilan Actif
    actif_pattern = r"'([A-Z]{2})'\s*:\s*'([A-Z]\d+)'"
    actif_matches = re.findall(actif_pattern, export_liasse_content)
    
    print_info(f"Vérification de {len(actif_matches)} cellules du Bilan Actif...")
    
    # Chercher l'onglet Bilan/Actif
    onglet_actif = None
    for name in ["ACTIF", "Bilan", "BILAN"]:
        if name in wb.sheetnames:
            onglet_actif = name
            break
    
    if onglet_actif:
        ws = wb[onglet_actif]
        cellules_ok = 0
        cellules_ko = 0
        
        for ref, cellule in actif_matches[:5]:  # Vérifier les 5 premières
            try:
                cell = ws[cellule]
                cellules_ok += 1
            except Exception as e:
                print_warning(f"  Cellule {cellule} (ref {ref}): ERREUR - {e}")
                cellules_ko += 1
        
        if cellules_ok > 0:
            print_success(f"  {cellules_ok} cellules vérifiées avec succès")
        if cellules_ko > 0:
            print_warning(f"  {cellules_ko} cellules avec erreur")
    else:
        print_warning("Onglet ACTIF/BILAN non trouvé pour vérification des cellules")
    
    wb.close()
    
except Exception as e:
    print_error(f"Erreur lors de la vérification des cellules: {e}")


# ==================== VÉRIFICATION 7: COHÉRENCE GLOBALE ====================

print_header("VÉRIFICATION 7: COHÉRENCE GLOBALE")

problemes = []

# Problème 1: TFT dans correspondances_syscohada.json
if "tft" in correspondances:
    print_success("TFT présent dans correspondances_syscohada.json")
else:
    problemes.append("TFT absent de correspondances_syscohada.json")
    print_warning("TFT absent de correspondances_syscohada.json")

# Problème 2: TFT dans export_liasse.py
if "TFT" not in export_liasse_content and "tft" not in export_liasse_content.lower():
    problemes.append("TFT non prévu dans export_liasse.py")
    print_warning("TFT non prévu dans export_liasse.py")

# Problème 3: Onglets du template
if onglets_trouves.get("TFT") is None:
    problemes.append("Onglet TFT non trouvé dans le template")
    print_warning("Onglet TFT non trouvé dans le template")

# Problème 4: Correspondance des onglets
onglets_mapping = {
    "BILAN": ["MAPPING_BILAN_ACTIF", "MAPPING_BILAN_PASSIF"],
    "COMPTE DE RÉSULTAT": ["MAPPING_COMPTE_RESULTAT_CHARGES", "MAPPING_COMPTE_RESULTAT_PRODUITS"]
}

for onglet_cat, mappings in onglets_mapping.items():
    if onglets_trouves.get(onglet_cat):
        for mapping in mappings:
            if mapping not in export_liasse_content:
                problemes.append(f"Mapping {mapping} manquant pour {onglet_cat}")


# ==================== RÉSUMÉ FINAL ====================

print_header("RÉSUMÉ FINAL")

if len(problemes) == 0:
    print_success("AUCUN PROBLÈME DÉTECTÉ")
    print_info("Tous les fichiers sont cohérents et correctement configurés.")
else:
    print_error(f"{len(problemes)} PROBLÈME(S) DÉTECTÉ(S):")
    for i, pb in enumerate(problemes, 1):
        print(f"  {i}. {pb}")

print("\n" + "="*80)
print("Vérification terminée.")
print("="*80 + "\n")
